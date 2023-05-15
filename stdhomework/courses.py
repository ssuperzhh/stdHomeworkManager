import os
from datetime import datetime

from flask import (
    Blueprint, flash, g, redirect, render_template, request, url_for, jsonify, current_app, send_file, json
)
from werkzeug.exceptions import abort
from werkzeug.utils import secure_filename

# from stdhomework.auth import login_required
from stdhomework.db import get_db

bp = Blueprint('courses', __name__, url_prefix='/courses')


# 根据 course_id 获取 course_notice 表中的数据
@bp.route('/course_notice/<int:course_id>', methods=['GET'])
def get_course_notice_by_course_id(course_id):
    db = get_db()
    # 创建一个 cursor 对象
    cursor = db.cursor()
    cursor.execute("SELECT A.*,B.course_name FROM course_noticeinfo A INNER JOIN courseinfo B "
                   "ON A.course_id=B.course_id WHERE A.course_id = %s", (course_id,))
    result = cursor.fetchall()
    if not result:
        return jsonify({'code': 404, 'msg': '未找到该课程下的通知'})
    return jsonify({"code": 0, 'msg': "", 'count': len(result), 'data': result})


# 向 course_notice 表中插入数据,添加课程公告
@bp.route('/course_notice', methods=['POST'])
def create_course_notice():
    # 解析请求中的 JSON 数据
    data = request.get_json()
    # print(data)
    # 从 JSON 数据中获取数据
    title = data['title']
    content = data['content']
    course_id = data['course_id']
    create_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    # 插入新数据
    db = get_db()
    try:
        # 创建一个 cursor 对象
        cursor = db.cursor()
        cursor.execute("INSERT INTO course_noticeinfo (title, content, create_time, course_id) VALUES (%s, %s, %s, %s)",
                       (title, content, create_time, course_id))
        db.commit()
        # 返回信息
        return jsonify({'code': 201, 'msg': '课程通知创建成功'})

    except Exception as e:
        # 如果出现异常，回滚并关闭游标
        db.rollback()
        cursor.close()

        # 返回错误信息
        return jsonify({'code': 500, 'msg': str(e)})


# 根据 id 更新 course_notice 表中的数据
@bp.route('/course_notice/update', methods=['PUT'])
def update_course_notice():
    # 解析请求中的 JSON 数据
    data = request.get_json()

    # 从 JSON 数据中获取数据
    id = data['id']
    title = data['title']
    content = data['content']
    create_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    course_id = data['course_id']

    # 更新数据
    db = get_db()
    # 创建一个 cursor 对象
    cursor = db.cursor()
    try:
        result = cursor.execute("UPDATE course_noticeinfo SET title = %s, content = %s, create_time = %s WHERE id = %s",
                                (title, content, create_time, id))
        db.commit()
        if result > 0:
            # 返回信息
            return jsonify({'code': 200, 'msg': '课程通知更新成功'})

    except Exception as e:
        # 如果出现异常，回滚并关闭游标
        db.rollback()
        cursor.close()
        # 返回错误信息
        return jsonify({'code': 500, 'msg': str(e)})


# 根据 id 删除 course_notice 表中的数据
@bp.route('/course_notice/delete', methods=['DELETE'])
def delete_course_notice_notice():
    id = request.json['id']
    # 删除数据
    db = get_db()
    # 创建一个 cursor 对象
    cursor = db.cursor()
    try:
        result = cursor.execute("DELETE FROM course_noticeinfo WHERE id = %s", (id,))
        db.commit()
        if result > 0:
            # 返回信息
            return jsonify({'code': 200, 'msg': '课程通知删除成功'})
        else:
            return jsonify({'code': 400, 'msg': '该课程不存在或已被删除'})

    except Exception as e:
        # 如果出现异常，回滚并关闭游标
        db.rollback()
        cursor.close()
        # 返回错误信息
        return jsonify({'code': 500, 'msg': str(e)})


# 获取当前课程下的作业
@bp.route('/homeworks/<int:course_id>', methods=['GET'])
def get_course_homeworks(course_id):
    db = get_db()
    try:
        sql = f'SELECT A.* FROM homeworkinfo A ' \
              f'WHERE course_id = "{course_id}"'
        cursor = db.cursor()
        cursor.execute(sql)
        db.commit()
        result = cursor.fetchall()
        if not result:
            return jsonify({'code': 404, 'msg': '未找到该课程下的作业'})
        return jsonify({"code": 0, 'msg': "", 'count': len(result), 'data': result})

    except Exception as e:
        # 如果出现异常，回滚并关闭游标
        db.rollback()
        cursor.close()

        # 返回错误信息
        return jsonify({'code': 500, 'msg': str(e)})


# 获取当前课程下的学生作业
@bp.route('/get_student_homeworks/<int:course_id>/<string:student_id>', methods=['GET'])
def get_student_homeworks(course_id, student_id):
    db = get_db()
    try:
        sql = f'SELECT A.*,B.state,B.score FROM homeworkinfo A ' \
              f'LEFT JOIN student_homework B ON A.homework_id=B.homework_id AND B.student_id="{student_id}"' \
              f'WHERE course_id = "{course_id}"'
        cursor = db.cursor()
        cursor.execute(sql)
        db.commit()
        result = cursor.fetchall()
        if not result:
            return jsonify({'code': 404, 'msg': '未找到该课程下的作业'})
        return jsonify({"code": 0, 'msg': "", 'count': len(result), 'data': result})

    except Exception as e:
        # 如果出现异常，回滚并关闭游标
        db.rollback()
        cursor.close()

        # 返回错误信息
        return jsonify({'code': 500, 'msg': str(e)})


# 添加课程作业
@bp.route('/homeworks', methods=['POST'])
def create_course_homeworks():
    data = request.get_json()
    homework_id = data['homework_id']
    homework_name = data['homework_name']
    type = data['type']
    course_id = data['course_id']
    create_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    # 插入新数据
    db = get_db()
    try:
        # 创建一个 cursor 对象
        cursor = db.cursor()
        cursor.execute(f"INSERT INTO homeworkinfo (homework_id, homework_name, date, type, course_id) "
                       f"VALUES ({homework_id}, '{homework_name}', '{create_time}', '{type}', {course_id})")
        db.commit()
        # 返回信息
        return jsonify({'code': 201, 'msg': '课程作业创建成功'})

    except Exception as e:
        # 如果出现异常，回滚并关闭游标
        db.rollback()
        cursor.close()

        # 返回错误信息
        return jsonify({'code': 500, 'msg': str(e)})


# 编辑课程作业
@bp.route('/homeworks/update', methods=['PUT'])
def update_course_homeworks():
    data = request.get_json()
    homework_id = data['homework_id']
    homework_name = data['homework_name']
    type = data['type']
    course_id = data['course_id']
    create_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # 更新数据
    db = get_db()
    # 创建一个 cursor 对象
    cursor = db.cursor()
    try:
        result = cursor.execute(
            "UPDATE homeworkinfo SET homework_name = %s, type = %s, date = %s WHERE homework_id = %s",
            (homework_name, type, create_time, homework_id))
        db.commit()
        if result > 0:
            # 返回信息
            return jsonify({'code': 200, 'msg': '课程作业更新成功'})

    except Exception as e:
        # 如果出现异常，回滚并关闭游标
        db.rollback()
        cursor.close()
        # 返回错误信息
        return jsonify({'code': 500, 'msg': str(e)})


# 删除某课程的作业
@bp.route('/homeworks/delete', methods=['DELETE'])
def delete_course_homeworks():
    homework_id = request.json['homework_id']
    # 删除数据
    db = get_db()
    # 创建一个 cursor 对象
    cursor = db.cursor()
    try:
        result = cursor.execute(f"DELETE FROM homeworkinfo WHERE homework_id = '{homework_id}'")
        db.commit()
        if result > 0:
            # 返回信息
            return jsonify({'code': 200, 'msg': '课程作业删除成功'})
        else:
            return jsonify({'code': 400, 'msg': '该课程作业不存在或已被删除'})

    except Exception as e:
        # 如果出现异常，回滚并关闭游标
        db.rollback()
        cursor.close()
        # 返回错误信息
        return jsonify({'code': 500, 'msg': str(e)})


# 创建允许上传的文件类型的集合
ALLOWED_EXTENSIONS = {'txt', 'xls', 'xlsx', 'xlsm', 'xlsb'}


def allowed_file(filename):
    return '.' in filename and \
        filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@bp.route('/upload', methods=['POST'])  # 中文文件上传尚未解决
def file_upload():
    # 检查上传的文件是否存在
    if 'file' not in request.files:
        return jsonify({'code': 400, 'msg': 'No file found'}), 400
    file = request.files['file']
    # 检查上传的文件名是否合法
    if file.filename == '':
        return jsonify({'code': 400, 'msg': 'No selected file'}), 400
    if not allowed_file(file.filename):
        return jsonify({'code': 400, 'msg': 'Invalid file type'}), 400
    # 保存文件
    filename = secure_filename(file.filename)
    homework_url = os.path.join(current_app.config['UPLOAD_FOLDER'], 'homework')
    file_url = os.path.join(homework_url, filename)
    file.save(file_url)  # 文件上传
    return jsonify({'code': 200, 'msg': 'File uploaded successfully'}), 200


@bp.route('/upload/add', methods=['POST'])
def file_upload_add():
    # 检查上传的文件是否存在
    if 'file' not in request.files:
        return jsonify({'code': 400, 'msg': 'No file found'}), 400
    file = request.files['file']
    # 检查上传的文件名是否合法
    if file.filename == '':
        return jsonify({'code': 400, 'msg': 'No selected file'}), 400
    if not allowed_file(file.filename):
        return jsonify({'code': 400, 'msg': 'Invalid file type'}), 400
    # 保存文件
    filename = secure_filename(file.filename)
    homework_url = os.path.join(current_app.config['UPLOAD_FOLDER'], 'homework')
    file_url = os.path.join(homework_url, filename)
    # file.save(file_url)  # 文件上传
    answer_file = request.files['answer_file']
    # 检查上传的文件名是否合法
    if answer_file.filename == '':
        return jsonify({'code': 400, 'msg': 'No selected file'}), 400
    if not allowed_file(answer_file.filename):
        return jsonify({'code': 400, 'msg': 'Invalid file type'}), 400
    # 保存文件
    answer_filename = secure_filename(answer_file.filename)
    homework_url = os.path.join(current_app.config['UPLOAD_FOLDER'], 'homework')
    answer_file_url = os.path.join(homework_url, answer_filename)
    # 获取表单中的参数
    form_data = request.form.get('form_data')
    form_data = json.loads(form_data)

    # 获取表单中的普通字段参数
    homework_id = form_data.get('homework_id')
    homework_name = form_data.get('homework_name')
    create_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    type = form_data.get('type')
    course_id = form_data.get('course_id')
    # 插入新数据
    db = get_db()
    try:
        # 创建一个 cursor 对象
        cursor = db.cursor()
        # 先添加作业
        cursor.execute(f"INSERT INTO homeworkinfo (homework_id, homework_name, date, type, course_id) "
                       f"VALUES ({homework_id}, '{homework_name}', '{create_time}', '{type}', {course_id})")
        # 再添加文件
        cursor.execute(f"INSERT INTO fileInfo (url, file_name, type, homework_id) "
                       f"VALUES ('{file_url}', '{filename}','{1}', {homework_id})",
                       )
        cursor.execute(f"INSERT INTO fileInfo (url, file_name, type, homework_id) "
                       f"VALUES ('{answer_file_url}', '{answer_filename}','{3}', {homework_id})",
                       )
        db.commit()
        cursor.close()
        return jsonify({'code': 200, 'msg': 'File uploaded successfully'}), 200
    except Exception as e:
        # 如果出现异常，回滚并关闭游标
        db.rollback()
        cursor.close()

        # 返回错误信息
        return jsonify({'code': 500, 'msg': str(e)})


@bp.route('/student/upload', methods=['POST'])  # 中文文件上传尚未解决
def student_file_upload():
    # 检查上传的文件是否存在
    if 'file' not in request.files:
        return jsonify({'code': 400, 'msg': 'No file found'}), 400
    file = request.files['file']
    # 检查上传的文件名是否合法
    if file.filename == '':
        return jsonify({'code': 400, 'msg': 'No selected file'}), 400
    if not allowed_file(file.filename):
        return jsonify({'code': 400, 'msg': 'Invalid file type'}), 400
    # 保存文件
    filename = secure_filename(file.filename)
    homework_url = os.path.join(current_app.config['UPLOAD_FOLDER'], 'homework')
    student_url = os.path.join(homework_url, 'student-answer')
    file_url = os.path.join(student_url, filename)
    file.save(file_url)  # 文件上传
    return jsonify({'code': 200, 'msg': 'File uploaded successfully'}), 200


@bp.route('/student/upload/add', methods=['POST'])
def student_file_upload_add():
    # 检查上传的文件是否存在
    if 'file' not in request.files:
        return jsonify({'code': 400, 'msg': 'No file found'}), 400
    file = request.files['file']
    # 检查上传的文件名是否合法
    if file.filename == '':
        return jsonify({'code': 400, 'msg': 'No selected file'}), 400
    if not allowed_file(file.filename):
        return jsonify({'code': 400, 'msg': 'Invalid file type'}), 400
    filename = secure_filename(file.filename)
    homework_url = os.path.join(current_app.config['UPLOAD_FOLDER'], 'homework')
    student_url = os.path.join(homework_url, 'student-answer')
    file_url = os.path.join(student_url, filename)
    # file.save(file_url)  # 文件上传
    # 获取表单中的参数
    form_data = request.form.get('form_data')
    form_data = json.loads(form_data)

    # 获取表单中的普通字段参数
    homework_id = form_data.get('homework_id')
    homework_name = form_data.get('homework_name')
    stu_id = form_data.get('stu_id')
    # 插入新数据
    db = get_db()
    try:
        # 创建一个 cursor 对象
        cursor = db.cursor()
        # 添加学生作业文件
        cursor.execute(f"INSERT INTO fileInfo (url, file_name, type, homework_id,stu_id) "
                       f"VALUES ('{file_url}', '{filename}','{2}', {homework_id},'{stu_id}')",  # 1作业文件，2学生作业，3答案
                       )
        cursor.execute(f"INSERT INTO student_homework (answer, state, homework_id,student_id) "
                       f"VALUES ('{file_url}', 'finished', {homework_id},'{stu_id}')",  # 1作业文件，2学生作业，3答案
                       )  # 添加进学生作业表中
        db.commit()
        cursor.close()
        return jsonify({'code': 200, 'msg': 'File uploaded successfully'}), 200
    except Exception as e:
        # 如果出现异常，回滚并关闭游标
        db.rollback()
        cursor.close()

        # 返回错误信息
        return jsonify({'code': 500, 'msg': str(e)})


# # 处理文件下载请求
# @bp.route('/download/<filename>', methods=['GET'])
# def file_download(filename):
#     homework_path = os.path.join(current_app.config['UPLOAD_FOLDER'], 'homework')
#     # 检查上传的文件名是否存在
#     file_path = os.path.abspath(os.path.join(homework_path, filename))
#     if not os.path.exists(file_path):
#         return jsonify({'code': 404, 'msg': 'File not found'}), 404
#     # 向浏览器发送文件
#     return send_file(file_path, as_attachment=True)
# 图片 MIME 类型映射表
MIME_TYPES = {
    '.txt': 'text/plain',
    '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
}


@bp.route('/download/<filename>', methods=['GET'])
def file_download(filename):
    homework_path = os.path.join(current_app.config['UPLOAD_FOLDER'], 'homework')
    file_path = os.path.abspath(os.path.join(homework_path, filename))

    if not os.path.isfile(file_path):
        return jsonify({'code': 404, 'msg': 'File not found.'}), 404

    # 从文件名获取文件扩展名
    ext = os.path.splitext(filename)[1]

    # 如果 MIME 类型未定义，使用通用二进制流
    mime_type = MIME_TYPES.get(ext, 'application/octet-stream')

    # 向浏览器发送文件
    return send_file(file_path, as_attachment=True, mimetype=mime_type)


@bp.route('/get_download_file/<int:homework_id>', methods=['GET'])
def get_download_file(homework_id):
    db = get_db()
    cursor = db.cursor()
    sql = f'SELECT DISTINCT file_name FROM fileinfo WHERE homework_id={homework_id}'
    try:
        cursor.execute(sql)
        file_name = cursor.fetchone()
        return jsonify({'code': 200, 'msg': '', 'data': file_name}), 200
    except Exception as e:
        # 返回错误信息
        return jsonify({'code': 500, 'msg': str(e)})


# excel作业自动批改
@bp.route('/auto_correct', methods=['GET'])
def auto_correct():
    import os
    import re
    from openpyxl import load_workbook
    from openpyxl import Workbook

    homework = os.path.join(current_app.config['UPLOAD_FOLDER'], 'homework')
    answer_file = os.path.join(homework, 'answer.xlsx')
    folder_path = os.path.join(homework, 'student-answer')
    check_file = os.path.join(homework, 'check.txt')
    result_file = os.path.join(homework, 'result.txt')
    score_file = os.path.join(homework, 'score.xlsx')
    # Load the answer Excel workbook
    answer_wb = load_workbook(answer_file, data_only=True)  # Set data_only to True to get calculated values

    # Read the check file
    with open(check_file, 'r', encoding='utf-8') as f:
        check_lines = f.readlines()

    def check_functions(formula, functions):
        if not functions:  # If the formula string is empty, return True
            return True

        for group in functions:
            for func in group:
                pattern = re.escape(func).replace(r'\*', '.*')
                if not re.search(pattern, formula, re.IGNORECASE):
                    break
            else:
                return True
        return False

    # Create a new Excel workbook for scores
    score_wb = Workbook()
    score_ws = score_wb.active
    score_ws.title = "Scores"
    score_ws.append(["Student File", "Total Score"])

    try:
        # Open the result file for writing
        with open(result_file, 'w') as f:
            # Iterate through each student file in the folder
            for student_file in os.listdir(folder_path):
                if student_file.endswith('.xlsx'):
                    student_wb = load_workbook(os.path.join(folder_path, student_file),
                                               data_only=True)  # Set data_only to True to get calculated values
                    student_wb_formulas = load_workbook(os.path.join(folder_path, student_file),
                                                        data_only=False)  # Set data_only to False to get formulas
                    total_score = 0
                    f.write(f'#####{student_file}:\n')

                    # Iterate through the check_lines
                    for line in check_lines:
                        line = line.strip()
                        if line.startswith('#'):  # Sheet name
                            sheet_name = line[1:]
                            f.write(f'{line}\n')  # Write the sheet name to the result file
                        else:  # Cell and scores
                            pattern = r'(\w+)-([\d.]+)-([\d.]+)(?:-{(.*?)})?'
                            match = re.match(pattern, line)
                            if match:
                                cell, value_score, formula_score, functions_str = match.groups()

                                value_score = float(value_score)
                                formula_score = float(formula_score)

                                # Extract function groups
                                if functions_str:
                                    functions = [group.split('and') for group in re.split(r'\s*;\s*', functions_str)]
                                else:
                                    functions = []

                                # Compare the student cell with the answer cell
                                student_cell = student_wb[sheet_name][cell]
                                student_cell_formula = student_wb_formulas[sheet_name][cell]
                                answer_cell = answer_wb[sheet_name][cell]

                                # Check the calculated value
                                if student_cell.value == answer_cell.value:
                                    f.write(f'{cell} value-{value_score:.1f} ')
                                    total_score += value_score

                                    # Check the formula if the calculated value matches
                                    if student_cell_formula.data_type == "f" and check_functions(
                                            student_cell_formula.value,
                                            functions):
                                        f.write(f'formula-{formula_score:.1f}\n')
                                        total_score += formula_score
                                    else:
                                        f.write(f'formula-0\n')
                                else:
                                    f.write(f'{cell} value-0 formula-0\n')

                    f.write(f'Total:{total_score:.1f}\n\n')
                    # Append the student file and total score to the score worksheet
                    score_ws.append([student_file, total_score])
        # Save the score workbook
        score_wb.save(score_file)

        # 取到分数存入学生作业表
        with open(result_file, 'r') as f:
            content = f.read()

        db = get_db()
        cursor = db.cursor()
        homework_id = request.args.get('homework_id')
        cursor.execute(f'SELECT file_name FROM fileinfo WHERE homework_id={homework_id} AND stu_id IS NULL ')
        filename = cursor.fetchone()['file_name']
        escaped_filename = re.escape(filename)
        pattern = r'###(\d+)-' + escaped_filename + r':\n[\s\S]*?Total:(\d+\.\d+)'
        matches = re.findall(pattern, content)
        for match in matches:
            student_id = match[0]
            total_score = float(match[1])
            sql = f'UPDATE student_homework SET score={total_score} WHERE homework_id={homework_id} AND student_id="{student_id}"'
            cursor.execute(sql)
        db.commit()
        cursor.close()
        return jsonify({'code': 200, 'msg': 'Results written to result.txt"'}), 200

    except Exception as e:
        # 返回错误信息
        return jsonify({'code': 500, 'msg': str(e)})

@bp.route('/get_lables', methods=['GET'])
def get_lables():
    lable_type = request.args.get('lable_type')
    db = get_db()
    cursor = db.cursor()
    sql = f"SELECT * FROM LABELS WHERE label_type = '{lable_type}'"
    cursor.execute(sql)
    lables_list = cursor.fetchall()
    return jsonify({'code': 0, 'msg': '', 'data': lables_list, 'count':len(lables_list)})

@bp.route('/normal_homework_add', methods=['POST'])
def normal_homework_add():
    # 获取前端通过 Ajax 提交的数据
    data = request.get_json()
    homework = data['homework']  # 作业基础数据
    create_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    choice = data['choice']  # 选择题
    fill = data['fill']  # 填空题
    answer_content = data['answer_content']  # 解答题
    homework_id = int(homework[0]['homework_id'])

    db = get_db()
    cursor = db.cursor()
    try:
        # 先添加作业
        cursor.execute(f"INSERT INTO homeworkinfo (homework_id, homework_name, date, type, course_id) "
                       f"VALUES({homework_id}, '{homework[0]['homework_name']}', '{create_time}', "
                       f"'{homework[0]['homework_type']}', {homework[0]['course_id']})")
        for i in range(len(choice)):
            answer = choice[i]['answer']
            knowledge = choice[i]['knowledge']
            level = choice[i]['level']
            question = choice[i]['question']
            question_score = choice[i]['score']
            A = choice[i]['options']['A']
            B = choice[i]['options']['B']
            C = choice[i]['options']['C']
            D = choice[i]['options']['D']
            # 添加选择题
            choice_sql = f'INSERT INTO questioninfo (question,a,b,c,d,truth,type,knowledge,level,homework_id,question_score) ' \
                         f'VALUES("{question}", "{A}", "{B}", "{C}", "{D}", "{answer}", 1, "{knowledge}", "{level}", {homework_id}, "{question_score}")'
            cursor.execute(choice_sql)

        for i in range(len(fill)):
            answer = fill[i]['answer']
            knowledge = fill[i]['knowledge']
            level = fill[i]['level']
            question = fill[i]['question']
            question_score = fill[i]['score']
            # 添加填空题
            fill_sql = f'INSERT INTO questioninfo (question,truth,type,knowledge,level,homework_id,question_score) ' \
                       f'VALUES("{question}","{answer}", 2, "{knowledge}", "{level}", {homework_id}, "{question_score}")'
            cursor.execute(fill_sql)

        for i in range(len(answer_content)):
            answer = answer_content[i]['answer']
            knowledge = answer_content[i]['knowledge']
            level = answer_content[i]['level']
            question = answer_content[i]['question']
            question_score = answer_content[i]['score']
            # 添加解答
            answer_content_sql = f'INSERT INTO questioninfo (question,truth,type,knowledge,level,homework_id,question_score) ' \
                                 f'VALUES ("{question}","{answer}", 3, "{knowledge}", "{level}", {homework_id}, "{question_score}")'
            cursor.execute(answer_content_sql)

        db.commit()
        cursor.close()
        return jsonify({'code': 201, 'msg': '数据提交成功'}), 200
    except Exception as e:
        # 如果出现异常，回滚并关闭游标
        db.rollback()
        cursor.close()
        # 返回错误信息
        return jsonify({'code': 500, 'msg': str(e)})


@bp.route('/get_normal_homework/<int:homework_id>', methods=['GET'])
def get_normal_homework(homework_id):
    db = get_db()
    cursor = db.cursor()

    try:
        choice_sql = f'SELECT * FROM questioninfo WHERE homework_id={homework_id} AND type=1'  # 获取选择题
        cursor.execute(choice_sql)
        choice_question = cursor.fetchall()
        fill_sql = f'SELECT * FROM questioninfo WHERE homework_id={homework_id} AND type=2'  # 获取填空题
        cursor.execute(fill_sql)
        fill_question = cursor.fetchall()
        content_sql = f'SELECT * FROM questioninfo WHERE homework_id={homework_id} AND type=3'  # 获取解答题
        cursor.execute(content_sql)
        content_question = cursor.fetchall()
        questions = {}
        questions['choice_question'] = choice_question
        questions['fill_question'] = fill_question
        questions['content_question'] = content_question
        return jsonify({'code': 0, 'msg': '', 'data': questions, 'count': len(questions)}), 200
    except Exception as e:
        # 如果出现异常，回滚并关闭游标
        db.rollback()
        cursor.close()
        # 返回错误信息
        return jsonify({'code': 500, 'msg': str(e)})


@bp.route('/student_question_add', methods=['POST'])
def student_question_add():
    # 获取前端通过 Ajax 提交的数据
    data = request.get_json()
    student_id = data['student_id']  # 作业基础数据
    homework_id = data['homework_id']  # 作业基础数据
    data.pop('student_id')
    data.pop('homework_id')
    db = get_db()
    cursor = db.cursor()
    try:
        for key, value in data.items():
            question_id = int(key.split('_')[1])
            sql = f'INSERT INTO student_question(answer,question_id,student_id,homework_id) VALUES (%s,%s,%s,%s)'
            values = (value, question_id, student_id, homework_id)
            cursor.execute(sql, values)

        cursor.execute(f"INSERT INTO student_homework (state, homework_id,student_id) "
                       f"VALUES ('finished', {homework_id},'{student_id}')",
                       )  # 添加进学生作业表中
        db.commit()
        cursor.close()
        return jsonify({'code': 200, 'msg': '数据提交成功'}), 200
    except Exception as e:
        # 如果出现异常，回滚并关闭游标
        db.rollback()
        cursor.close()
        # 返回错误信息
        return jsonify({'code': 500, 'msg': str(e)})
